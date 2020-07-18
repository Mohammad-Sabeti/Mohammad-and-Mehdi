from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws1=wb.create_sheet("sheet1")
ws1['A1'] = "mahdi"
ws1['B1'] = "09029582464"
ws1['A2'] = "mohamad"
ws1['B2'] = "09021523612"
ws1['A3'] = "ali"
ws1['B3'] = "09108763210"
ws1['A4'] = "saeed"
ws1['B4'] = "09902451188"
ws1['A5'] = "omid"
ws1['B5'] = ""
ws1['A6'] = "emad"
ws1['B6'] = "09029581716"
ws1['A7'] = "mohamad"
ws1['B7'] = "09027654321"
ws1['A8'] = ""
ws1['B8'] = "09336139966"
ws1['A9'] = "sajad"
ws1['B9'] = "09144497900"
ws1['A10'] = "mobin"
ws1['B10'] = "09336485490"
wb.save("sample.xlsx")


from openpyxl import load_workbook
wb = load_workbook('sample.xlsx')
sheet_range=wb['sheet1']
name=["A","1"]
phone=["B","1"]

def searchname(m):
    wname=input(" enter name:")
    x=0
    i = 1
    while i>0:
        n1 = name[0] + name[1]
        p1 = phone[0] + phone[1]
        if sheet_range[n1].value==None:
            while m<100:
                i += 1
                I = str(i)
                name[1] = I
                n1 = name[0] + name[1]
                if sheet_range[n1].value !=None:
                    break
                m+=1
            if m==100:
                break
        if wname==sheet_range[n1].value:
            x+=1
            if sheet_range[p1].value==None:
                print("this name have not a phone number")
            else:
                print('phone:', sheet_range[p1].value)
        i += 1
        I = str(i)
        name[1] = I
        phone[1] = I
    if x==0:
        print("not found")

def searchphone(n):
    wphone=input(" enter phone:")
    y=0
    j = 1
    name[1]='1'
    phone[1]='1'
    while j>0:
        n2 = name[0] + name[1]
        p2 = phone[0] + phone[1]
        if sheet_range[p2].value==None:
            while n<100:
                j += 1
                J = str(j)
                phone[1] = J
                p2 = phone[0] + phone[1]
                if sheet_range[p2].value !=None:
                    break
                n+=1
            if n==100:
                break
        if wphone==sheet_range[p2].value:
            y+=1
            if sheet_range[n2].value==None:
                print("this phonenumber have not a name")
            else:
                print('name:', sheet_range[n2].value)
                break
        j += 1
        J = str(j)
        name[1] = J
        phone[1] = J
    if y==0:
        print("not found")

searchname(0)
searchphone(0)





